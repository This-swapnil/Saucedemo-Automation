import os
import openpyxl
from platformdirs import user_data_dir


class UserCreds:
    user_credentials = [
        {"username": "standard_user", "password": "secret_sauce"},
        {"username": "standard_user", "password": "secr_sauce"},
        {"username": "stdard_user", "password": "secret_sauce"},
        {"username": "standard_er", "password": "secret_sce"},
        {"username": "locked_out_user", "password": "secret_sauce"},
        {"username": "problem_user", "password": "secret_sauce"},
        {"username": "performance_glitch_user", "password": "secret_sauce"},
    ]

    @staticmethod
    def get_user_creds_data():
        file_path = os.path.join(
            "testData",
            "usercredentials.xlsx",
        )
        work_book = openpyxl.load_workbook(file_path)
        sheet = work_book["all"]
        user_data = []
        for row in sheet.iter_rows(min_row=2, max_col=3):
            creds_dict = {"username": row[1].value, "password": row[2].value}
            user_data.append(creds_dict)
        return user_data

    @staticmethod
    def get_invalid_user_credentials():
        file_path = os.path.join(
            "testData",
            "usercredentials.xlsx",
        )
        work_book = openpyxl.load_workbook(file_path)
        sheet = work_book["invalid"]
        user_data = []
        for row in sheet.iter_rows(min_row=2, max_col=2):
            cred_dict = {"username": row[0].value, "password": row[1].value}
            user_data.append(cred_dict)
        return user_data
